import openpyxl
import os

input_wb = openpyxl.load_workbook('produceSales.xlsx')
input_sheet = input_wb.active
output_wb = openpyxl.Workbook()
output_sheet = output_wb.active
col_1 = []
col_2 = []

for row in range(1, input_sheet.max_row + 1):
    for col in range(1, input_sheet.max_column + 1):
        if col == 1:
            col_val_1 = input_sheet.cell(row=row, column=col).value
            col_1.append(col_val_1)
        else:
            col_val_2 = input_sheet.cell(row=row, column=col).value
            col_2.append(col_val_2)

col_count = 1
col_count_2 = 1
for item in col_1:
    output_sheet.cell(row=1, column=col_count).value = item
    col_count += 1
for item in col_2:
    output_sheet.cell(row=2, column=col_count_2).value = item
    col_count_2 += 1
os.chdir(r'C:\Users\Ozgur2\Desktop')
output_wb.save('post_cell_invert.xlsx')
