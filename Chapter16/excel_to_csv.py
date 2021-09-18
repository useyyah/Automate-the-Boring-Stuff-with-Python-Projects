import csv
import openpyxl
import os
import re
import shutil

os.chdir(r"C:\Users\Ozgur2\Desktop\example_spreadsheets")

path = os.getcwd()

os.makedirs("csv_files", exist_ok=True)
new_path = path + "\\" + "csv_files"

for excel_file in os.listdir("."):
    if not excel_file.endswith("xlsx"):
        continue
    workbook = openpyxl.load_workbook(excel_file)

    for sheets in workbook.sheetnames:
        wb_name = re.sub(".xlsx", "", excel_file)
        csv_name = wb_name + "_" + sheets + ".csv"
        csv_file = open(csv_name, "w", newline="")
        csv_writer = csv.writer(csv_file)
        sheet = workbook.active

        for row_num in range(1, sheet.max_row + 1):
            row_data = []
            for col_num in range(1, sheet.max_column + 1):
                cell_data = sheet.cell(row=row_num, column=col_num).value
                row_data.append(cell_data)
            csv_writer.writerow(row_data)
        csv_file.close()
        shutil.move(os.path.join(path, csv_name), os.path.join(new_path, csv_name))
print("Finished")
